# -+-coding=utf8-+-
import os
import sys

import chardet
import openpyxl
import xlrd
import zipfile

from email.mime.multipart import MIMEMultipart
from email.utils import COMMASPACE
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr
import chardet
import email
import os
import base64

import shutil


class Mail:
    def __init__(self):
        self.attachment = []
        self.attachmentname = []
        self._sender = False
        self._subject = False
        self._content = False

    def set_sender(self, nickname, addr):
        self.nickname = nickname
        self.addr = addr
        self._sender = True

    def set_subject(self, subject):
        self.subject = subject
        self._subject = True

    def set_content(self, content):
        self.content_txt = content
        self._content = True

    def add_attachment(self, attachmentFile: str or list):
        if not isinstance(attachmentFile, list):
            attachmentFile = [attachmentFile]
        for eachfile in attachmentFile:
            basename = os.path.basename(eachfile)
            with open(eachfile, 'rb') as f:
                part = MIMEApplication(f.read())
                filename = base64.b64encode(basename.encode('utf-8'))
                part.add_header('Content-Disposition', 'attachment',
                                filename='=?UTF-8?B?' + filename.decode() + '?=')
            self.attachment.append(part)
            self.attachmentname.append(basename)

    def make_mail(self):
        data = MIMEMultipart()
        if self._sender and self._subject and self._content:
            self.sender = formataddr(
                (Header(self.nickname, 'utf-8').encode(), self.addr))
            data['From'] = self.sender
            data['To'] = COMMASPACE.join([])
            data['Subject'] = Header(self.subject, 'utf-8').encode()
            try:
                self.content = MIMEText(self.content_txt, 'plain', 'utf-8')
            except UnicodeEncodeError:
                self.content = MIMEText(self.content_txt, 'plain', 'gb2312')
            data.attach(self.content)
            for each in self.attachment:
                data.attach(each)
            return data
        return None


def judge_pure_english(keyword):
    return all(ord(c) < 128 for c in keyword)


class excel:
    def __init__(self):
        self._addrs = []
        self._last = 0
        self._email = []
        self.double = []
        self.error = []

    @staticmethod
    def is_email(x):
        if isinstance(x, str):
            slt = x.split('@')
            return len(slt) == 2 and ('.' in slt[1])
        return False

    @property
    def emails(self):
        if self._last != len(self._addrs):
            st = set()
            self._email = []
            for each in self._addrs:
                if each not in st:
                    st.add(each)
                    if judge_pure_english(each[0]):
                        self._email.append(each)
                    else:
                        self.error.append(each)
            else:
                self.double.append(each)
            self._last = len(self._addrs)
        return [x[0] for x in self._email]

    def handle_xlsx(self, ws):
        idx = None
        for i, cell in enumerate(next(ws.rows), 1):
            if 'address' in cell.value.lower() or 'email' in cell.value.lower() or '邮箱' in cell.value:
                idx = i
                break
        if not idx:
            return ()
        emails = (each.value for rows in ws.iter_rows(
            min_row=2, min_col=idx, max_col=idx) for each in rows)
        return (each.strip() for each in emails if self.is_email(each))

    def handle_xls(self, ws):
        idx = None
        for i, cell in enumerate(ws.row_values(0)):
            if 'address' in cell.lower() or 'email' in cell.lower() or '邮箱' in cell:
                idx = i
                break
        if not idx:
            return ()
        emails = (each for each in ws.col_values(idx))
        return (each.strip() for each in emails if self.is_email(each))

    def add(self, fname, colnums=None):
        real_cols = []
        addr = []

        if isinstance(colnums, str):
            slt = colnums.split(',')
            for rng in slt:
                if '-' in rng:
                    ft = rng.split('-')
                    if len(ft) != 2:
                        return False
                    f, t = list(map(int, ft))
                    real_cols += list(range(f-1, t))
                else:
                    real_cols.append(int(rng)-1)
        elif isinstance(colnums, (list, tuple)):
            real_cols = [x-1 for x in colnums if isinstance(x, int)]

        if '.xlsx' in fname:
            wb = openpyxl.load_workbook(fname)

            if not real_cols:
                rets = list(self.handle_xlsx(wb.active))
                rets = zip(rets, [fname]*len(rets))
                addr += rets
            else:
                sname = wb.sheetnames
                for index in real_cols:
                    try:
                        ws = wb.get_sheet_by_name(sname[index])
                    except:
                        continue
                    rets = list(self.handle_xlsx(ws))
                    rets = zip(rets, [fname]*len(rets))
                    addr += rets

        elif '.xls' in fname:
            wb = xlrd.open_workbook(fname)
            if not real_cols:
                rets = list(self.handle_xlsx(wb.sheet_by_index(0)))
                rets = zip(rets, [fname]*len(rets))
                addr += rets
            if not real_cols:
                real_cols = list(range(wb.nsheets))
            for index in real_cols:
                try:
                    ws = wb.sheet_by_index(index)
                except:
                    continue
                # addr += list(self.handle_xls(ws))
                rets = list(self.handle_xls(ws))
                rets = zip(rets, [fname]*len(rets))
                addr += rets

        self._addrs += addr
        sys.stdout.write('success add {} column: {}\n'.format(
            os.path.basename(fname), ' '.join(map(lambda x: str(x+1), real_cols)) if real_cols else 'all'))
        # print('success add {} column: {}'.format(
        #     os.path.basename(fname), ' '.join(map(lambda x: str(x+1), real_cols)) if real_cols else 'all'))


def get(fname):
    shutil.rmtree('./temp', ignore_errors=True)
    zf = zipfile.ZipFile(fname)
    zf.extractall('./temp')
    dirlist = os.listdir('./temp')
    if 'config.txt' in dirlist:
        basedir = './temp'
    else:
        for fn in dirlist:
            fn = os.path.join('./temp', fn)
            if os.path.isdir(fn):
                if 'config.txt' in os.listdir(fn):
                    basedir = fn
                    break

    with open(os.path.join(basedir, 'config.txt'), 'rb') as fp:
        file_data = fp.read()
        file_encoding = chardet.detect(file_data)
        configfile = file_data.decode(file_encoding['encoding'])

    print(configfile)

    for line in configfile.split('\n'):
        if len(line.strip()) <= 1:
            continue
        slt = line.split(':')
        if len(slt) != 2:
            return
        if slt[0] == 'emails':
            emails = excel()
            emails.add(os.path.join(basedir, slt[1].strip()))
            emails = emails.emails
        elif slt[0] == 'nickname':
            nickname = slt[1].strip()
        elif slt[0] == 'title':
            title = slt[1].strip()
        elif slt[0] == 'debug':
            dbg = True if ('1' in slt[1].strip(
            ) or 't' in slt[1].strip().lower()) else False
        elif slt[0] == 'password':
            password = slt[1].strip()
        else:
            return
    loc = locals().keys()
    # print(loc)
    if not (('password' in loc) and ('emails' in loc) and ('nickname' in loc) and ('title' in loc) and ('dbg' in loc)):
        return None

    if 'content.txt' in os.listdir(basedir):
        with open(os.path.join(basedir, 'content.txt'), 'rb') as fp:
            file_data = fp.read()
            file_encoding = chardet.detect(file_data)
            content = file_data.decode(file_encoding['encoding'])

    attachmentDir = os.path.join(basedir, 'attachments')
    attachments = list(map(lambda x: os.path.join(
        attachmentDir, x), os.listdir(attachmentDir)))

    mail = Mail()
    mail.set_sender(nickname, 'dian@hust.edu.cn')
    mail.set_subject(title)
    mail.set_content(content)
    mail.add_attachment(attachments)

    shutil.rmtree('./temp', ignore_errors=True)
    return (emails, password, nickname, title, attachments, content, mail, dbg)


if __name__ == "__main__":
    print(get('test.zip'))
